"""Business rules for the stock catalog."""

from __future__ import annotations

import uuid
import csv
import io
from io import BytesIO
from typing import Optional

from openpyxl import load_workbook

from app.models.stock import StockItem, utcnow
from app.repositories.stock import StockRepository
from app.schemas.stock import StockCreate, StockUpdate
from app.schemas.stock_import import (
    StockImportError,
    StockImportResult,
)


class StockNotFoundError(Exception):
    def __init__(self, item_id: uuid.UUID) -> None:
        self.item_id = item_id
        super().__init__(f"stock item {item_id} not found")


def _clean_aliases(aliases: list[str]) -> list[str]:
    return list(
        dict.fromkeys(
            alias.strip()
            for alias in aliases
            if alias.strip()
        )
    )


class StockService:

    def __init__(self, repository: StockRepository) -> None:
        self._repository = repository

    def list_stock(
        self,
        business_id: uuid.UUID,
    ) -> list[StockItem]:
        return self._repository.list_active(
            business_id=business_id
        )

    def get_active(
        self,
        item_id: uuid.UUID,
        business_id: uuid.UUID,
    ) -> Optional[StockItem]:

        item = self._repository.get(item_id)

        if (
            item is None
            or not item.is_active
            or item.business_id != business_id
        ):
            return None

        return item

    def get_for_business(
        self,
        item_id: uuid.UUID,
        business_id: uuid.UUID,
    ) -> Optional[StockItem]:

        item = self._repository.get(item_id)

        if (
            item is None
            or item.business_id != business_id
        ):
            return None

        return item

    def create_stock(
        self,
        data: StockCreate,
        business_id: uuid.UUID,
    ) -> StockItem:

        item = StockItem(
            id=uuid.uuid4(),
            business_id=business_id,
            name=data.name,
            sku=data.sku,
            unit=data.unit,
            unit_price=data.unit_price,
            quantity_available=data.quantity_available,
            low_stock_threshold=data.low_stock_threshold,
            aliases=_clean_aliases(data.aliases),
            created_at=utcnow(),
            updated_at=utcnow(),
        )

        return self._repository.add(item)

    def update_stock(
        self,
        item_id: uuid.UUID,
        data: StockUpdate,
        business_id: uuid.UUID,
    ) -> StockItem:

        existing = self._repository.get(item_id)

        if (
            existing is None
            or not existing.is_active
            or existing.business_id != business_id
        ):
            raise StockNotFoundError(item_id)

        updates = data.model_dump(exclude_unset=True)

        if updates.get("aliases") is not None:
            updates["aliases"] = _clean_aliases(
                updates["aliases"]
            )

        updated = existing.model_copy(
            update={
                **updates,
                "updated_at": utcnow(),
            }
        )

        return self._repository.update(updated)

    def delete_stock(
        self,
        item_id: uuid.UUID,
        business_id: uuid.UUID,
    ) -> None:

        existing = self._repository.get(item_id)

        if (
            existing is None
            or not existing.is_active
            or existing.business_id != business_id
        ):
            raise StockNotFoundError(item_id)

        self._repository.soft_delete(item_id)

    def import_csv(
        self,
        content: bytes,
        business_id: uuid.UUID,
    ) -> StockImportResult:

        errors: list[StockImportError] = []
        imported = 0
        skipped = 0

        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            return StockImportResult(
                imported=0,
                skipped=0,
                failed=1,
                errors=[
                    StockImportError(
                        row=1,
                        message="CSV file must be UTF-8 encoded.",
                    )
                ],
            )

        reader = csv.DictReader(io.StringIO(text))

        required_columns = {
            "name",
            "unit",
            "unit_price",
            "quantity_available",
        }

        actual_columns = set(
            reader.fieldnames or []
        )

        missing_columns = (
            required_columns - actual_columns
        )

        if missing_columns:
            return StockImportResult(
                imported=0,
                skipped=0,
                failed=1,
                errors=[
                    StockImportError(
                        row=1,
                        message=(
                            "Missing required columns: "
                            + ", ".join(
                                sorted(missing_columns)
                            )
                        ),
                    )
                ],
            )

        existing_items = self.list_stock(
            business_id=business_id
        )

        existing_names = {
            item.name.strip().lower()
            for item in existing_items
        }

        existing_skus = {
            item.sku.strip().lower()
            for item in existing_items
            if item.sku
        }

        seen_names: set[str] = set()
        seen_skus: set[str] = set()

        for row_number, row in enumerate(
            reader,
            start=2,
        ):
            try:
                name = (
                    row.get("name") or ""
                ).strip()

                sku = (
                    row.get("sku") or ""
                ).strip() or None

                unit = (
                    row.get("unit") or ""
                ).strip()

                unit_price_raw = (
                    row.get("unit_price") or ""
                ).strip()

                quantity_raw = (
                    row.get("quantity_available") or ""
                ).strip()

                threshold_raw = (
                    row.get("low_stock_threshold")
                    or "0"
                ).strip()

                aliases_raw = (
                    row.get("aliases") or ""
                ).strip()

                if not name:
                    raise ValueError(
                        "name is required"
                    )

                if not unit:
                    raise ValueError(
                        "unit is required"
                    )

                try:
                    unit_price = float(
                        unit_price_raw
                    )
                except ValueError:
                    raise ValueError(
                        "unit_price must be a number"
                    )

                try:
                    quantity_available = float(
                        quantity_raw
                    )
                except ValueError:
                    raise ValueError(
                        "quantity_available must be a number"
                    )

                try:
                    low_stock_threshold = float(
                        threshold_raw
                    )
                except ValueError:
                    raise ValueError(
                        "low_stock_threshold must be a number"
                    )

                if unit_price < 0:
                    raise ValueError(
                        "unit_price cannot be negative"
                    )

                if quantity_available < 0:
                    raise ValueError(
                        "quantity_available cannot be negative"
                    )

                if low_stock_threshold < 0:
                    raise ValueError(
                        "low_stock_threshold cannot be negative"
                    )

                name_key = name.lower()
                sku_key = (
                    sku.lower()
                    if sku
                    else None
                )

                if name_key in existing_names:
                    skipped += 1
                    continue

                if (
                    sku_key
                    and sku_key in existing_skus
                ):
                    skipped += 1
                    continue

                if name_key in seen_names:
                    skipped += 1
                    continue

                if (
                    sku_key
                    and sku_key in seen_skus
                ):
                    skipped += 1
                    continue

                aliases = [
                    alias.strip()
                    for alias in aliases_raw.split(",")
                    if alias.strip()
                ]

                stock_data = StockCreate(
                    name=name,
                    sku=sku,
                    unit=unit,
                    unit_price=unit_price,
                    quantity_available=quantity_available,
                    low_stock_threshold=low_stock_threshold,
                    aliases=aliases,
                )

                self.create_stock(
                    stock_data,
                    business_id=business_id,
                )

                imported += 1
                seen_names.add(name_key)

                if sku_key:
                    seen_skus.add(sku_key)

            except Exception as exc:
                errors.append(
                    StockImportError(
                        row=row_number,
                        message=str(exc),
                    )
                )

        return StockImportResult(
            imported=imported,
            skipped=skipped,
            failed=len(errors),
            errors=errors,
        )

    def import_xlsx(
        self,
        file_content: bytes,
        business_id: uuid.UUID,
    ) -> StockImportResult:

        errors: list[StockImportError] = []
        imported = 0
        skipped = 0

        try:
            workbook = load_workbook(
                filename=BytesIO(file_content),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            return StockImportResult(
                imported=0,
                skipped=0,
                failed=1,
                errors=[
                    StockImportError(
                        row=1,
                        message=f"Unable to read Excel file: {exc}",
                    )
                ],
            )

        try:
            worksheet = workbook.active

            rows = list(
                worksheet.iter_rows(
                    values_only=True
                )
            )

            if not rows:
                return StockImportResult(
                    imported=0,
                    skipped=0,
                    failed=1,
                    errors=[
                        StockImportError(
                            row=1,
                            message="Excel file is empty.",
                        )
                    ],
                )

            headers = [
                str(value).strip()
                if value is not None
                else ""
                for value in rows[0]
            ]

            required_columns = {
                "name",
                "unit",
                "unit_price",
                "quantity_available",
            }

            actual_columns = set(headers)

            missing_columns = (
                required_columns - actual_columns
            )

            if missing_columns:
                return StockImportResult(
                    imported=0,
                    skipped=0,
                    failed=1,
                    errors=[
                        StockImportError(
                            row=1,
                            message=(
                                "Missing required columns: "
                                + ", ".join(
                                    sorted(missing_columns)
                                )
                            ),
                        )
                    ],
                )

            existing_items = self.list_stock(
                business_id=business_id
            )

            existing_names = {
                item.name.strip().lower()
                for item in existing_items
            }

            existing_skus = {
                item.sku.strip().lower()
                for item in existing_items
                if item.sku
            }

            seen_names: set[str] = set()
            seen_skus: set[str] = set()

            for row_number, values in enumerate(
                rows[1:],
                start=2,
            ):
                try:
                    row = {
                        headers[index]: (
                            values[index]
                            if index < len(values)
                            else None
                        )
                        for index in range(
                            len(headers)
                        )
                    }

                    name = str(
                        row.get("name") or ""
                    ).strip()

                    sku_value = row.get("sku")

                    sku = (
                        str(sku_value).strip()
                        if sku_value is not None
                        else None
                    )

                    unit = str(
                        row.get("unit") or ""
                    ).strip()

                    if not name:
                        raise ValueError(
                            "name is required"
                        )

                    if not unit:
                        raise ValueError(
                            "unit is required"
                        )

                    try:
                        unit_price = float(
                            row.get("unit_price")
                        )
                    except (TypeError, ValueError):
                        raise ValueError(
                            "unit_price must be a number"
                        )

                    try:
                        quantity_available = float(
                            row.get(
                                "quantity_available"
                            )
                        )
                    except (TypeError, ValueError):
                        raise ValueError(
                            "quantity_available must be a number"
                        )

                    threshold_value = row.get(
                        "low_stock_threshold"
                    )

                    try:
                        low_stock_threshold = float(
                            threshold_value
                            if threshold_value not in (
                                None,
                                "",
                            )
                            else 0
                        )
                    except (TypeError, ValueError):
                        raise ValueError(
                            "low_stock_threshold must be a number"
                        )

                    if unit_price < 0:
                        raise ValueError(
                            "unit_price cannot be negative"
                        )

                    if quantity_available < 0:
                        raise ValueError(
                            "quantity_available cannot be negative"
                        )

                    if low_stock_threshold < 0:
                        raise ValueError(
                            "low_stock_threshold cannot be negative"
                        )

                    name_key = name.lower()

                    sku_key = (
                        sku.lower()
                        if sku
                        else None
                    )

                    # Existing database item
                    if name_key in existing_names:
                        skipped += 1
                        continue

                    if (
                        sku_key
                        and sku_key in existing_skus
                    ):
                        skipped += 1
                        continue

                    # Duplicate inside current Excel file
                    if name_key in seen_names:
                        skipped += 1
                        continue

                    if (
                        sku_key
                        and sku_key in seen_skus
                    ):
                        skipped += 1
                        continue

                    aliases_value = row.get(
                        "aliases"
                    )

                    aliases = []

                    if aliases_value is not None:
                        aliases = [
                            alias.strip()
                            for alias in str(
                                aliases_value
                            ).split(",")
                            if alias.strip()
                        ]

                    stock_data = StockCreate(
                        name=name,
                        sku=sku,
                        unit=unit,
                        unit_price=unit_price,
                        quantity_available=quantity_available,
                        low_stock_threshold=low_stock_threshold,
                        aliases=aliases,
                    )

                    self.create_stock(
                        stock_data,
                        business_id=business_id,
                    )

                    imported += 1
                    seen_names.add(name_key)

                    if sku_key:
                        seen_skus.add(sku_key)

                except Exception as exc:
                    errors.append(
                        StockImportError(
                            row=row_number,
                            message=str(exc),
                        )
                    )

            return StockImportResult(
                imported=imported,
                skipped=skipped,
                failed=len(errors),
                errors=errors,
            )

        finally:
            workbook.close()


    def restore_stock_for_cancellation(
        self,
        item_id: uuid.UUID,
        quantity: float,
        business_id: uuid.UUID,
    ) -> tuple[StockItem, float, float]:
        """
        Restore stock quantity when a sale/invoice is cancelled.

        This intentionally allows soft-deleted stock items because
        historical invoices may reference products that were deleted later.
        The stock item remains deleted; only its quantity is restored.
        """

        existing = self._repository.get(item_id)

        if (
            existing is None
            or existing.business_id != business_id
        ):
            raise StockNotFoundError(item_id)

        quantity_before = existing.quantity_available
        quantity_after = quantity_before + quantity

        updated = existing.model_copy(
            update={
                "quantity_available": quantity_after,
                "updated_at": utcnow(),
            }
        )

        self._repository.update(updated)

        return updated, quantity_before, quantity_after